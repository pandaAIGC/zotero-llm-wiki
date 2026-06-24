# -*- coding: utf-8 -*-
r"""
Zotero LLM Wiki - 批量入库脚本（MinerU 批量并行解析）

流水线:
  Phase 1: 下载/复制 PDF + 检查缓存
  Phase 2: 批量提交 MinerU（并行解析）
  Phase 3: 收集结果 → 切块 → Embedding → ChromaDB

用法:
  cd <path-to-zotero-llm-wiki>
  .venv\Scripts\python.exe run_ingest.py                    # 全量入库
  .venv\Scripts\python.exe run_ingest.py --incremental      # 只入库新增
  .venv\Scripts\python.exe run_ingest.py --limit 10         # 最多处理10篇
  .venv\Scripts\python.exe run_ingest.py --collection 示例主题   # 只处理指定 Collection
  .venv\Scripts\python.exe run_ingest.py --batch-size 10    # 每批提交10篇到MinerU
"""
import sys
import io
import os
import logging
import json
import re
import time
import hashlib
from datetime import datetime, timezone
from pathlib import Path

# Ensure UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

# Ensure working directory is correct
os.chdir(Path(__file__).parent)

# Ensure data directories exist
Path("parsed").mkdir(exist_ok=True)
Path("data/chroma_db").mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

import httpx

import config
import zotero_sync
import pdf_parser
import chunker
import vector_store

# Number of files per MinerU batch submission
MINERU_BATCH_SIZE = 20
# Polling timeout per batch (seconds)
MINERU_POLL_TIMEOUT = 1800
MINERU_NO_PROGRESS_TIMEOUT = 900

_DEFAULT_JCR_FILE_ENV = os.environ.get("ZOTERO_JCR_FILE", "")
DEFAULT_JCR_FILE = Path(_DEFAULT_JCR_FILE_ENV) if _DEFAULT_JCR_FILE_ENV else None
DAILY_STATE_FILE = config.DATA_DIR / "daily_incremental_state.json"


def _parse_zotero_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    text = value.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(text)
    except ValueError:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _iso_utc(dt: datetime) -> str:
    return dt.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def _daily_since_baseline() -> datetime:
    if DAILY_STATE_FILE.exists():
        try:
            state = json.loads(DAILY_STATE_FILE.read_text(encoding="utf-8"))
            dt = _parse_zotero_datetime(str(state.get("last_run_started_at") or ""))
            if dt:
                return dt
        except Exception as exc:
            logger.warning("Could not read daily incremental state %s: %s", DAILY_STATE_FILE, exc)

    stats_path = config.DATA_DIR / "last_ingest_stats.json"
    if stats_path.exists():
        return datetime.fromtimestamp(stats_path.stat().st_mtime, tz=timezone.utc)
    return datetime.now(timezone.utc)


def _save_daily_state(run_started_at: datetime, stats: dict) -> None:
    DAILY_STATE_FILE.write_text(
        json.dumps(
            {
                "last_run_started_at": _iso_utc(run_started_at),
                "last_stats": {
                    "total": stats.get("total"),
                    "success": stats.get("success"),
                    "chunks": stats.get("chunks"),
                    "api_limited": stats.get("api_limited"),
                    "no_actionable": stats.get("no_actionable"),
                },
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def _save_stats(stats: dict, parse_failures: dict | None = None) -> Path:
    stats_path = config.DATA_DIR / "last_ingest_stats.json"
    if parse_failures is not None:
        _save_parse_failures(parse_failures)
    with open(stats_path, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    return stats_path


def _is_api_limit_error(exc: BaseException) -> bool:
    status = None
    if isinstance(exc, httpx.HTTPStatusError):
        status = exc.response.status_code
    else:
        status = getattr(getattr(exc, "response", None), "status_code", None)

    if status in {402, 429}:
        return True

    text = str(exc).lower()
    markers = [
        "429",
        "too many requests",
        "rate limit",
        "quota",
        "exceeded",
        "insufficient",
        "resource package",
        "余额不足",
        "无可用资源包",
        "额度",
        "限流",
        "频率",
    ]
    return any(marker in text for marker in markers)


def _load_parse_failures() -> dict[str, dict]:
    path = config.PARSE_FAILURES_FILE
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("Could not read parse failure registry %s: %s", path, exc)
        return {}
    if not isinstance(data, dict):
        return {}
    return {str(k): v for k, v in data.items() if isinstance(v, dict)}


def _save_parse_failures(parse_failures: dict[str, dict]) -> None:
    config.PARSE_FAILURES_FILE.write_text(
        json.dumps(parse_failures, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _record_parse_failure(parse_failures: dict[str, dict], key: str, reason: str) -> None:
    entry = parse_failures.get(key, {})
    entry["attempts"] = int(entry.get("attempts") or 0) + 1
    entry["reason"] = reason
    entry["updated"] = time.strftime("%Y-%m-%dT%H:%M:%S")
    parse_failures[key] = entry


def _clear_parse_failure(parse_failures: dict[str, dict], key: str) -> None:
    parse_failures.pop(key, None)


def _parse_failure_attempts(parse_failures: dict[str, dict], key: str) -> int:
    return int((parse_failures.get(key) or {}).get("attempts") or 0)


def _parse_failure_skip_reason(parse_failures: dict[str, dict], key: str) -> str | None:
    max_attempts = int(getattr(config, "PARSE_FAILURE_MAX_ATTEMPTS", 0) or 0)
    attempts = _parse_failure_attempts(parse_failures, key)
    if max_attempts > 0 and attempts >= max_attempts:
        reason = (parse_failures.get(key) or {}).get("reason") or "parse_failed"
        return f"{reason} after {attempts} attempt(s)"
    return None


def _load_suspect_pdf_skip_keys() -> set[str]:
    path = config.SUSPECT_PDF_SKIP_KEYS_FILE
    if not path.exists():
        return set()
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return set()
    if isinstance(data, list):
        return {str(k) for k in data if k}
    if isinstance(data, dict):
        keys = data.get("keys", [])
        return {str(k) for k in keys if k}
    return set()


def _save_suspect_pdf_skip_keys(keys: set[str]) -> None:
    config.SUSPECT_PDF_SKIP_KEYS_FILE.write_text(
        json.dumps(sorted(keys), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _normalize_journal(value: str) -> str:
    text = str(value or "").lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_title(value: str) -> str:
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalize_doi(value: str) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"^(?:https?://(?:dx\.)?doi\.org/|doi:\s*)", "", text)
    return text.rstrip(" .")


def _title_similarity(a: str, b: str) -> float:
    a_tokens = set(_normalize_title(a).split())
    b_tokens = set(_normalize_title(b).split())
    if not a_tokens or not b_tokens:
        return 0.0
    return len(a_tokens & b_tokens) / len(a_tokens | b_tokens)


_PDF_TITLE_STOPWORDS = {
    "the", "and", "for", "with", "from", "into", "via", "using", "based",
    "analysis", "study", "novel", "role", "between", "through", "reveals",
    "of", "in", "on", "to", "a", "an", "as", "by", "or", "is", "are",
}


def _zotero_item_summary(data: dict, col_map: dict[str, str]) -> dict:
    """Convert Zotero item data to the compact shape used by the ingest pipeline."""
    authors = []
    for creator in data.get("creators", []):
        if creator.get("creatorType") == "author":
            last = creator.get("lastName", "")
            first = creator.get("firstName", "")
            name = f"{last} {first}".strip()
            if name:
                authors.append(name)

    date_str = data.get("date", "")
    year = None
    if date_str:
        try:
            year = int(date_str[:4])
        except (ValueError, IndexError):
            pass

    collection_names = []
    for col_key in data.get("collections", []):
        col_name = col_map.get(col_key, "")
        if col_name:
            collection_names.append(col_name)
    if not collection_names:
        collection_names = [config.DEFAULT_COLLECTION]

    return {
        "key": data["key"],
        "title": data.get("title", ""),
        "authors": authors,
        "year": year,
        "doi": data.get("DOI", ""),
        "item_type": data["itemType"],
        "journal": data.get("publicationTitle", ""),
        "journal_abbreviation": data.get("journalAbbreviation", ""),
        "issn": data.get("ISSN", ""),
        "url": data.get("url", ""),
        "abstract": data.get("abstractNote", ""),
        "collection_names": collection_names,
        "has_pdf": False,
    }


def _fetch_items_by_keys(zot, item_keys: set[str]) -> list[dict]:
    col_map = {col["key"]: col["name"] for col in zotero_sync.list_collections(zot)}
    items = []
    for key in sorted(item_keys):
        try:
            data = zot.item(key)["data"]
        except Exception as exc:
            logger.error("  %s: failed to fetch Zotero item metadata: %s", key, exc)
            continue
        if data.get("itemType") in {"attachment", "note", "annotation"}:
            continue
        items.append(_zotero_item_summary(data, col_map))
    return items


def _title_tokens(value: str) -> list[str]:
    tokens = re.findall(r"[a-zA-Z0-9]{4,}", str(value or "").lower())
    out: list[str] = []
    for token in tokens:
        if token in _PDF_TITLE_STOPWORDS:
            continue
        if token not in out:
            out.append(token)
    return out[:18]


def _pdf_matches_item(pdf_path: Path, item: dict, cache: dict[tuple[str, str], dict]) -> dict:
    """Conservatively validate whether a PDF appears to match a Zotero parent item.

    This is only used to break suspicious same-hash ties. If validation cannot
    confidently prove a match, the caller should keep treating the attachment as
    suspicious instead of ingesting potentially wrong full text.
    """
    title = str(item.get("title") or "")
    doi = _normalize_doi(item.get("doi", "") or item.get("DOI", ""))
    cache_key = (str(pdf_path.resolve()), _normalize_title(title))
    if cache_key in cache:
        return cache[cache_key]

    result = {
        "ok": False,
        "reason": "not_checked",
        "title_score": 0.0,
        "title_token_hits": 0,
        "title_token_total": 0,
    }
    tokens = _title_tokens(title)
    result["title_token_total"] = len(tokens)
    if not pdf_path.exists():
        result["reason"] = "pdf_missing"
        cache[cache_key] = result
        return result
    if len(tokens) < 4:
        result["reason"] = "not_enough_title_tokens"
        cache[cache_key] = result
        return result

    try:
        from pypdf import PdfReader

        reader = PdfReader(str(pdf_path))
        pieces: list[str] = []
        for page in reader.pages[:4]:
            try:
                pieces.append(page.extract_text() or "")
            except Exception:
                continue
    except Exception as exc:
        result["reason"] = f"pdf_read_failed:{type(exc).__name__}"
        cache[cache_key] = result
        return result

    text = "\n".join(pieces).lower()
    if len(text.strip()) < 200:
        result["reason"] = "pdf_text_unavailable"
        cache[cache_key] = result
        return result

    hits = sum(1 for token in tokens if token in text)
    score = hits / max(1, len(tokens))
    result["title_token_hits"] = hits
    result["title_score"] = round(score, 3)

    filename_has_doi = bool(doi and _normalize_doi(pdf_path.name).find(doi.replace("/", "_")) >= 0)
    text_has_doi = bool(doi and doi in text)
    if score >= 0.35 or filename_has_doi or text_has_doi:
        result["ok"] = True
        result["reason"] = "accepted_by_title_or_doi"
    else:
        result["reason"] = "title_mismatch_rejected"

    cache[cache_key] = result
    return result


def _local_parse_pdf_text(pdf_path: Path, item: dict) -> tuple[str | None, str]:
    """Extract readable PDF text locally as a conservative MinerU fallback."""
    if not config.LOCAL_PARSE_FALLBACK:
        return None, "disabled"
    if not pdf_path.exists():
        return None, "pdf_missing"

    try:
        import fitz  # PyMuPDF

        doc = fitz.open(str(pdf_path))
        page_count = doc.page_count
        if page_count > config.LOCAL_PARSE_FALLBACK_MAX_PAGES:
            doc.close()
            return None, f"too_many_pages:{page_count}"

        pages: list[str] = []
        for page_index in range(page_count):
            text = doc.load_page(page_index).get_text("text") or ""
            if text.strip():
                pages.append(f"## Page {page_index + 1}\n\n{text.strip()}")
        doc.close()
    except Exception as exc:
        return None, f"local_read_failed:{type(exc).__name__}"

    body = "\n\n".join(pages).strip()
    if len(body) < config.LOCAL_PARSE_FALLBACK_MIN_CHARS:
        return None, f"too_little_text:{len(body)}"

    if item.get("item_type") == "webpage":
        match = _pdf_matches_item(pdf_path, item, {})
        if not match.get("ok"):
            return None, f"webpage_title_mismatch:{match.get('reason')}"

    title = item.get("title") or pdf_path.stem
    markdown = (
        f"# {title}\n\n"
        "> Parsed locally from PDF text after MinerU returned an empty or failed result. "
        "Tables, formulas, and figures may be less structured than MinerU output.\n\n"
        f"{body}\n"
    )
    return markdown, "local_text"


def _try_local_parse_fallback(item: dict, pdf_path: Path, parsed: dict[str, tuple]) -> bool:
    key = item["key"]
    md, reason = _local_parse_pdf_text(pdf_path, item)
    if not md:
        logger.warning("  %s: local parse fallback unavailable (%s)", key, reason)
        return False

    cache_dir = config.PARSED_DIR / key
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_md = cache_dir / f"{key}.md"
    cache_md.write_text(md, encoding="utf-8")
    parsed[key] = (item, md)
    logger.info("  %s: local parse fallback succeeded (%s chars)", key, len(md))
    return True


def _same_pdf_is_expected(first_item: dict, item: dict) -> tuple[bool, str]:
    first_doi = _normalize_doi(first_item.get("doi", "") or first_item.get("DOI", ""))
    this_doi = _normalize_doi(item.get("doi", "") or item.get("DOI", ""))
    if first_doi and this_doi and first_doi == this_doi:
        return True, "same DOI"

    first_title = first_item.get("title", "")
    this_title = item.get("title", "")
    if _normalize_title(first_title) == _normalize_title(this_title):
        return True, "same title"
    if _title_similarity(first_title, this_title) >= 0.90:
        return True, "near-identical title"
    return False, ""


def _file_sha256(path: Path, cache: dict[Path, str]) -> str:
    resolved = path.resolve()
    cached = cache.get(resolved)
    if cached:
        return cached
    h = hashlib.sha256()
    with resolved.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    digest = h.hexdigest()
    cache[resolved] = digest
    return digest


def _load_jcr_table(path: Path) -> dict[str, dict]:
    """Load JCR rows keyed by normalized journal name/abbreviation/ISSN.

    openpyxl is imported lazily so normal ingestion does not need it.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for --high-impact-only. "
            "Install it in the Zotero LLM Wiki venv with: .\\.venv\\Scripts\\python.exe -m pip install openpyxl"
        ) from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    rows = ws.iter_rows(values_only=True)
    for row in rows:
        values = [str(c or "").strip() for c in row]
        if "Journal name" in values and "2024 JIF" in values:
            header = values
            break
    if header is None:
        raise RuntimeError(f"JCR header not found in {path}")

    col = {name: idx for idx, name in enumerate(header)}
    index: dict[str, dict] = {}

    def _value(row, name):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx] or "").strip()

    def _jif(row) -> float:
        raw = _value(row, "2024 JIF")
        try:
            return float(raw)
        except ValueError:
            return 0.0

    def _add(key: str, record: dict):
        norm = _normalize_journal(key)
        if not norm:
            return
        old = index.get(norm)
        if old is None or record["jif"] > old["jif"]:
            index[norm] = record

    for row in rows:
        journal = _value(row, "Journal name")
        if not journal:
            continue
        record = {
            "journal": journal,
            "abbreviation": _value(row, "JCR Abbreviation"),
            "issn": _value(row, "ISSN"),
            "eissn": _value(row, "eISSN"),
            "category": _value(row, "Category"),
            "jif": _jif(row),
            "quartile": _value(row, "JIF Quartile"),
            "rank": _value(row, "JIF Rank"),
            "five_year_jif": _value(row, "5 Year JIF"),
        }
        _add(record["journal"], record)
        _add(record["abbreviation"], record)
        _add(record["issn"], record)
        _add(record["eissn"], record)

    return index


def _jcr_match(item: dict, jcr_index: dict[str, dict]) -> dict | None:
    candidates = [
        item.get("journal", ""),
        item.get("journal_abbreviation", ""),
        item.get("issn", ""),
    ]
    for candidate in candidates:
        record = jcr_index.get(_normalize_journal(candidate))
        if record:
            return record
    return None


def _filter_high_impact(
    items: list[dict],
    jcr_file: Path,
    min_impact_factor: float,
    require_q1: bool,
    keep_unknown_journal: bool,
) -> list[dict]:
    jcr_index = _load_jcr_table(jcr_file)
    kept = []
    excluded_low = 0
    excluded_unknown = 0
    for item in items:
        record = _jcr_match(item, jcr_index)
        if record is None:
            if keep_unknown_journal:
                item["jcr"] = None
                kept.append(item)
            else:
                excluded_unknown += 1
            continue
        passed_if = record["jif"] >= min_impact_factor
        passed_q = (not require_q1) or (str(record.get("quartile", "")).upper() == "Q1")
        if passed_if and passed_q:
            item["jcr"] = record
            kept.append(item)
        else:
            excluded_low += 1
    print(
        f"  高质量期刊过滤: 保留 {len(kept)} 篇；"
        f"排除低IF/非Q1 {excluded_low} 篇；排除未知期刊 {excluded_unknown} 篇"
    )
    return kept


def _get_ingested_collections_by_key() -> dict[str, set[str]]:
    """Return paper key -> display collection names already present in ChromaDB."""
    keys: dict[str, set[str]] = {}
    try:
        for col in vector_store.list_collections():
            col_name = col["name"]
            for key in vector_store.get_paper_keys(col_name):
                keys.setdefault(key, set()).add(col_name)
    except Exception:
        pass
    return keys


def _is_fully_ingested(item: dict, ingested_by_key: dict[str, set[str]]) -> bool:
    """Treat an item as ingested only when every target collection has its chunks."""
    key = item["key"]
    present = ingested_by_key.get(key, set())
    if not present:
        return False
    target_collections = _target_collections(item)
    return all(col_name in present for col_name in target_collections)


def _target_collections(item: dict) -> list[str]:
    """Return target collection names with stable order and no duplicates."""
    names = item.get("collection_names") or [config.DEFAULT_COLLECTION]
    unique: list[str] = []
    seen: set[str] = set()
    for name in names:
        col_name = name or config.DEFAULT_COLLECTION
        if col_name in seen:
            continue
        seen.add(col_name)
        unique.append(col_name)
    return unique or [config.DEFAULT_COLLECTION]


def _cached_parse_path(item_key: str, pdf_path: Path) -> Path | None:
    """Return an existing parse cache path, preferring the stable item-key name."""
    candidates = [
        config.PARSED_DIR / item_key / f"{item_key}.md",
        config.PARSED_DIR / item_key / f"{pdf_path.stem}.md",
    ]
    return next((p for p in candidates if p.exists()), None)


def _has_cached_parse(item_key: str, pdf_path: Path) -> bool:
    """检查是否有缓存的解析结果（key 命名优先，向后兼容 stem 命名）"""
    return _cached_parse_path(item_key, pdf_path) is not None


def _ensure_collection_mapping(col_name: str) -> None:
    """确保 Collection 有中英文映射，没有则自动生成 slug 并注册"""
    if col_name == config.DEFAULT_COLLECTION:
        return
    if col_name in config._NAME_MAP:
        return  # already mapped
    # Auto-generate slug: use pinyin-style hash for Chinese names
    import hashlib
    h = hashlib.md5(col_name.encode()).hexdigest()[:10]
    slug = f"col-{h}"
    config.register_collection_mapping(col_name, slug)
    logger.info(f"  自动注册映射: '{col_name}' → '{slug}'")


def _download_pdf_with_retries(item_key: str, attempts: int = 3) -> Path | None:
    """Resolve/download a Zotero PDF without letting one transient failure abort a full run."""
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            return zotero_sync.download_pdf(item_key=item_key)
        except Exception as exc:
            if _is_api_limit_error(exc):
                raise
            last_error = exc
            if attempt < attempts:
                wait = min(5 * attempt, 15)
                logger.warning(
                    "  %s: PDF lookup failed (%s/%s): %s; retrying in %ss",
                    item_key,
                    attempt,
                    attempts,
                    exc,
                    wait,
                )
                time.sleep(wait)
    logger.error("  %s: PDF lookup failed after %s attempts: %s - SKIP", item_key, attempts, last_error)
    return None


def _process_paper(item: dict, markdown_text: str) -> int:
    """切块 + Embedding + 入库，返回 chunk 数量"""
    key, chunks, target_collections = _prepare_paper_chunks(item, markdown_text)
    return _process_prepared_paper(key, chunks, target_collections)


def _process_prepared_paper(key: str, chunks: list, target_collections: list[str]) -> int:
    """Embedding + Chroma upsert for chunks prepared by _prepare_paper_chunks."""
    if not chunks:
        logger.warning(f"  {key}: no chunks")
        return 0

    logger.info(
        "  %s: processing %s chunks across %s collection(s): %s",
        key,
        len(chunks),
        len(target_collections),
        ", ".join(target_collections),
    )
    # Ensure all target collections have mappings
    for col_name in target_collections:
        _ensure_collection_mapping(col_name)

    total = 0
    for col_name in target_collections:
        n = vector_store.add_chunks(chunks, collection_name=col_name)
        total += n
        logger.info(f"  {key}: {n} chunks -> [{col_name}]")
    return total


def _prepare_paper_chunks(item: dict, markdown_text: str) -> tuple[str, list, list[str]]:
    """Prepare chunks without calling embedding, so budget checks are cheap."""
    key = item["key"]
    title = item.get("title", "?")
    paper_metadata = {
        "key": key,
        "title": title,
        "authors": ", ".join(item.get("authors", [])),
        "year": str(item.get("year", "")),
        "doi": item.get("doi", ""),
        "url": item.get("url", ""),
        "abstract": item.get("abstract", ""),
    }
    chunks = chunker.chunk_markdown(markdown_text, paper_metadata=paper_metadata)
    target_collections = _target_collections(item)
    return key, chunks, target_collections


def _submit_and_wait_batch(
    client: httpx.Client,
    pdf_paths: list[Path],
    item_keys: list[str],
) -> dict[str, str | None]:
    """
    Batch-submit PDFs to MinerU via raw httpx REST API.
    Returns {item_key: markdown_text | None}.
    """
    opts = {
        "model": config.MINERU_MODEL,
        "formula": True,
        "table": True,
        "language": "ch",
    }

    logger.info(f"  Submitting {len(pdf_paths)} PDFs to MinerU...")
    path_strs = [str(p) for p in pdf_paths]
    batch_id = pdf_parser._get_upload_urls(client, path_strs, opts, ocr=True, pages=None)
    logger.info(f"  Batch ID: {batch_id}")

    # Poll until all done/failed
    deadline = time.monotonic() + MINERU_POLL_TIMEOUT
    interval = 5.0
    results_map: dict[str, str | None] = {}
    last_done_count = -1
    last_progress_at = time.monotonic()

    while True:
        resp = client.get(f"{pdf_parser._API_BASE}/extract-results/batch/{batch_id}")
        resp.raise_for_status()
        body = resp.json()
        if body.get("code", body.get("success", True)) not in (0, True):
            raise RuntimeError(f"MinerU polling error: {body.get('msg', body.get('message', 'unknown'))}")

        results = body["data"]["extract_result"]
        done_count = sum(1 for r in results if r.get("state") in ("done", "failed"))
        total = len(results)
        logger.info(f"  Polling: {done_count}/{total} done")
        if done_count > last_done_count:
            last_done_count = done_count
            last_progress_at = time.monotonic()

        if done_count == total:
            # Download markdown + images for each result
            outputs = []
            for r in results:
                if r.get("state") != "done":
                    outputs.append(("", []))
                    continue
                try:
                    out = pdf_parser._download_results(client, [r])
                    outputs.append(out[0] if out else ("", []))
                except Exception as e:
                    logger.warning(f"  Download failed: {e}")
                    outputs.append(("", []))

            for i, (md_text, images) in enumerate(outputs):
                item_key = item_keys[i]
                if md_text:
                    cache_dir = config.PARSED_DIR / item_key
                    cache_dir.mkdir(parents=True, exist_ok=True)
                    cache_md = cache_dir / f"{item_key}.md"
                    cache_md.write_text(md_text, encoding="utf-8")

                    if images:
                        images_dir = cache_dir / "images"
                        images_dir.mkdir(parents=True, exist_ok=True)
                        for img_name, img_data in images:
                            (images_dir / img_name).write_bytes(img_data)

                    results_map[item_key] = md_text
                    logger.info(f"  [{item_key}] done: {len(md_text)} chars")
                else:
                    logger.warning(f"  [{item_key}] failed or empty")
                    results_map[item_key] = None
            break

        if time.monotonic() > deadline:
            raise TimeoutError(f"MinerU batch timeout ({MINERU_POLL_TIMEOUT}s)")

        if time.monotonic() - last_progress_at > MINERU_NO_PROGRESS_TIMEOUT:
            raise TimeoutError(f"MinerU batch no progress for {MINERU_NO_PROGRESS_TIMEOUT}s")

        time.sleep(min(interval, max(0, deadline - time.monotonic())))
        interval = min(interval * 1.5, 30.0)

    return results_map


def run(
    incremental: bool = True,
    limit: int = 0,
    force_parse: bool = False,
    collection_filter: str | None = None,
    mineru_batch_size: int = MINERU_BATCH_SIZE,
    high_impact_only: bool = False,
    jcr_file: Path | None = None,
    min_impact_factor: float = 7.0,
    require_q1: bool = True,
    keep_unknown_journal: bool = False,
    dry_run: bool = False,
    max_parse_papers: int = 0,
    max_embed_papers: int = 0,
    max_embed_chunks: int = 0,
    item_keys: set[str] | None = None,
    local_parse_only: bool = False,
    daily_new_only: bool = False,
):
    """运行完整入库管线"""
    run_started_at = datetime.now(timezone.utc)
    print("=" * 60)
    print("  Zotero LLM Wiki - 批量入库管线")
    print("=" * 60)
    print(f"  本地 Zotero:     {config.ZOTERO_LOCAL_STORAGE}")
    print(f"  ChromaDB:        {config.CHROMA_DIR}")
    print(f"  Parsed:          {config.PARSED_DIR}")
    print(f"  Embedding:       {config.EMBED_PROVIDER}:{config.EMBED_MODEL}")
    print(f"  MinerU batch:    {mineru_batch_size}")
    if item_keys:
        print(f"  Item keys:       {', '.join(sorted(item_keys))}")
    if daily_new_only:
        print("  Daily mode:      only Zotero items added after previous daily run")
    if local_parse_only:
        print("  Parse mode:      local PDF text only (MinerU disabled)")
    if max_parse_papers > 0:
        print(f"  Parse budget:    max {max_parse_papers} papers")
    if max_embed_papers > 0:
        print(f"  Embed budget:    max {max_embed_papers} papers")
    if max_embed_chunks > 0:
        print(f"  Embed budget:    max {max_embed_chunks} chunk writes")
    if high_impact_only:
        print(f"  High-impact:     JIF >= {min_impact_factor}" + (" and Q1" if require_q1 else ""))
    print()

    # -- Phase 0: Fetch paper list --
    zot = zotero_sync._get_client()
    if item_keys:
        items = _fetch_items_by_keys(zot, item_keys)
    else:
        items = zotero_sync.list_items(zot, check_pdf=False)
    print(f"  Zotero 论文总数: {len(items)}")

    daily_since = None
    if daily_new_only and not item_keys:
        daily_since = _daily_since_baseline()
        before = len(items)
        items = [
            item for item in items
            if (dt := _parse_zotero_datetime(item.get("date_added"))) is not None and dt > daily_since
        ]
        print(f"  每日新增过滤: {before} -> {len(items)} 篇 (dateAdded > {_iso_utc(daily_since)})")

    if collection_filter:
        items = [
            it for it in items
            if any(
                collection_filter.lower() in name.lower()
                for name in it.get("collection_names", [])
            )
        ]
        print(f"  过滤 '{collection_filter}': {len(items)} 篇")

    if incremental:
        already = _get_ingested_collections_by_key()
        before = len(items)
        items = [it for it in items if not _is_fully_ingested(it, already)]
        print(f"  增量模式: {before - len(items)} 已入库, {len(items)} 待处理")

    if high_impact_only:
        jcr_path = jcr_file or DEFAULT_JCR_FILE
        if jcr_path is None:
            raise SystemExit("JCR file is required when --high-impact-only is set. Use --jcr-file or ZOTERO_JCR_FILE.")
        if not jcr_path.exists():
            raise SystemExit(f"JCR file not found: {jcr_path}")
        before = len(items)
        items = _filter_high_impact(
            items,
            jcr_path,
            min_impact_factor,
            require_q1,
            keep_unknown_journal,
        )
        print(f"  JCR 文件: {jcr_path}")
        print(f"  高质量过滤: {before} -> {len(items)} 篇")

    if limit > 0:
        items = items[:limit]
        print(f"  限制: 最多处理 {limit} 篇")

    if not items:
        print("\n  没有需要处理的论文")
        if daily_new_only and not dry_run:
            stats = {
                "total": 0,
                "daily_new_only": True,
                "daily_since": _iso_utc(daily_since or run_started_at),
                "high_impact_only": high_impact_only,
                "min_impact_factor": min_impact_factor if high_impact_only else None,
                "require_q1": require_q1 if high_impact_only else None,
                "cached": 0,
                "parsed": 0,
                "parse_failed_empty": 0,
                "parse_failed_empty_keys": [],
                "no_pdf_skipped": 0,
                "parse_failures_skipped": 0,
                "parse_failure_max_attempts": config.PARSE_FAILURE_MAX_ATTEMPTS,
                "success": 0,
                "skipped": 0,
                "failed": 0,
                "chunks": 0,
                "api_limited": False,
                "parse_mode": "local" if local_parse_only else "mineru",
                "embed_provider": config.EMBED_PROVIDER,
                "embed_model": config.EMBED_MODEL,
                "parse_budget_stopped": False,
                "max_parse_papers": max_parse_papers or None,
                "embed_budget_stopped": False,
                "max_embed_papers": max_embed_papers or None,
                "max_embed_chunks": max_embed_chunks or None,
                "suspect_pdf_duplicates": 0,
                "exact_pdf_duplicate_skipped": 0,
                "no_actionable": True,
            }
            stats_path = _save_stats(stats)
            _save_daily_state(run_started_at, stats)
            print(f"\nStats saved: {stats_path}")
        return

    if dry_run:
        print("\n  Dry-run: 以下文献会进入解析/向量化候选，不会调用 MinerU 或 Embedding")
        for i, item in enumerate(items, 1):
            jcr = item.get("jcr") or {}
            journal = item.get("journal") or item.get("journal_abbreviation") or "unknown journal"
            jcr_text = (
                f"JIF={jcr.get('jif')} {jcr.get('quartile')} | {jcr.get('journal')}"
                if jcr else "JCR=unknown"
            )
            print(f"  [{i}] {item.get('key')} | {item.get('year')} | {journal} | {jcr_text}")
            print(f"      {item.get('title', '')[:140]}")
        return

    print()

    parse_failures = _load_parse_failures()
    parse_failures_skipped = 0
    no_pdf_skipped = 0

    # -- Phase 1: Download PDF + Check cache --
    print("=" * 60)
    print("  Phase 1: 下载 PDF + 检查缓存")
    print("=" * 60)

    cached: dict[str, tuple] = {}       # item_key -> (item, markdown_text)
    need_parse: list[tuple] = []        # [(item, pdf_path), ...]
    pdf_hash_cache: dict[Path, str] = {}
    pdf_match_cache: dict[tuple[str, str], dict] = {}
    seen_pdf_hashes: dict[str, tuple[dict, Path]] = {}
    suspect_pdf_duplicate_keys: set[str] = set()
    persistent_suspect_pdf_skip_keys = _load_suspect_pdf_skip_keys()
    suspect_pdf_duplicate_records: list[dict] = []
    exact_pdf_duplicate_skipped = 0

    for i, item in enumerate(items, 1):
        key = item["key"]
        title = item.get("title", "?")[:50]
        if key in persistent_suspect_pdf_skip_keys:
            suspect_pdf_duplicate_keys.add(key)
            suspect_pdf_duplicate_records.append({
                "key": key,
                "title": item.get("title", ""),
                "action": "skip_persistent_suspect",
            })
            logger.error(
                "  [%s/%s] %s: persistent suspicious PDF skip - SKIP",
                i,
                len(items),
                key,
            )
            continue

        # Download PDF (prefer local copy)
        try:
            pdf_path = _download_pdf_with_retries(key)
        except Exception as e:
            if _is_api_limit_error(e):
                logger.error(f"  API limit reached during PDF lookup; stopping Phase 1: {e}")
                break
            logger.error(f"  [{i}/{len(items)}] {key}: PDF lookup crashed: {e} - SKIP", exc_info=True)
            pdf_path = None
        if pdf_path is None:
            no_pdf_skipped += 1
            logger.warning(f"  [{i}/{len(items)}] {key}: no PDF - SKIP")
            continue

        try:
            pdf_hash = _file_sha256(pdf_path, pdf_hash_cache)
        except Exception as exc:
            logger.error(f"  [{i}/{len(items)}] {key}: cannot hash PDF ({exc}) - SKIP")
            continue

        item["_pdf_hash"] = pdf_hash
        item["_pdf_path"] = str(pdf_path)
        seen = seen_pdf_hashes.get(pdf_hash)
        if seen:
            first_item, first_pdf_path = seen
            same_pdf_expected, same_pdf_reason = _same_pdf_is_expected(first_item, item)
            if same_pdf_expected:
                exact_pdf_duplicate_skipped += 1
                logger.warning(
                    "  [%s/%s] %s: duplicate PDF hash of %s with %s - SKIP duplicate",
                    i,
                    len(items),
                    key,
                    first_item["key"],
                    same_pdf_reason,
                )
                continue

            first_match = _pdf_matches_item(first_pdf_path, first_item, pdf_match_cache)
            this_match = _pdf_matches_item(pdf_path, item, pdf_match_cache)
            if first_match.get("ok") and not this_match.get("ok"):
                suspect_pdf_duplicate_keys.add(key)
                suspect_action = "skip_duplicate_only"
                logger.error(
                    "  [%s/%s] %s: suspicious duplicate PDF hash shared with %s; duplicate title mismatch - SKIP duplicate",
                    i,
                    len(items),
                    key,
                    first_item["key"],
                )
            elif this_match.get("ok") and not first_match.get("ok"):
                suspect_pdf_duplicate_keys.add(first_item["key"])
                seen_pdf_hashes[pdf_hash] = (item, pdf_path)
                suspect_action = "skip_first_keep_duplicate"
                logger.error(
                    "  [%s/%s] %s: suspicious duplicate PDF hash shared with %s; first title mismatch - KEEP current",
                    i,
                    len(items),
                    key,
                    first_item["key"],
                )
            else:
                suspect_pdf_duplicate_keys.add(first_item["key"])
                suspect_pdf_duplicate_keys.add(key)
                suspect_action = "skip_both"
                logger.error(
                    "  [%s/%s] %s: suspicious duplicate PDF hash shared with %s but titles differ - SKIP both",
                    i,
                    len(items),
                    key,
                    first_item["key"],
                )
            suspect_pdf_duplicate_records.append({
                "pdf_hash": pdf_hash,
                "first_key": first_item["key"],
                "first_title": first_item.get("title", ""),
                "first_pdf_path": str(first_pdf_path),
                "first_pdf_match": first_match,
                "duplicate_key": key,
                "duplicate_title": item.get("title", ""),
                "duplicate_pdf_path": str(pdf_path),
                "duplicate_pdf_match": this_match,
                "action": suspect_action,
            })
            if key in suspect_pdf_duplicate_keys:
                continue

        seen_pdf_hashes[pdf_hash] = (item, pdf_path)

        # Check cache
        if not force_parse and _has_cached_parse(key, pdf_path):
            cache_md = _cached_parse_path(key, pdf_path)
            if cache_md is None:
                logger.warning(f"  [{i}/{len(items)}] {key}: cache disappeared - need parse")
                need_parse.append((item, pdf_path))
                continue
            md = cache_md.read_text(encoding="utf-8")
            if len(md.strip()) >= config.MIN_PARSED_CACHE_CHARS:
                cached[key] = (item, md)
                logger.info(f"  [{i}/{len(items)}] {key}: cached ({len(md)} chars)")
                continue
            logger.warning(
                "  [%s/%s] %s: cache too short (%s chars) - need parse",
                i,
                len(items),
                key,
                len(md.strip()),
            )

        skip_reason = None if force_parse else _parse_failure_skip_reason(parse_failures, key)
        if skip_reason:
            parse_failures_skipped += 1
            logger.warning(
                "  [%s/%s] %s: previous parse failure (%s) - SKIP",
                i,
                len(items),
                key,
                skip_reason,
            )
            continue

        need_parse.append((item, pdf_path))
        logger.info(f"  [{i}/{len(items)}] {key}: {title} -> need parse ({pdf_path.stat().st_size / 1024 / 1024:.1f} MB)")

    print(f"\n  缓存命中: {len(cached)} 篇")
    print(f"  需要解析: {len(need_parse)} 篇")
    if no_pdf_skipped:
        print(f"  无 PDF 跳过: {no_pdf_skipped} 篇")
    if parse_failures_skipped:
        print(f"  解析失败历史跳过: {parse_failures_skipped} 篇")
    if suspect_pdf_duplicate_keys:
        cached = {
            key: value
            for key, value in cached.items()
            if key not in suspect_pdf_duplicate_keys
        }
        need_parse = [
            (item, pdf_path)
            for item, pdf_path in need_parse
            if item["key"] not in suspect_pdf_duplicate_keys
        ]
        report_path = config.DATA_DIR / "suspect_pdf_duplicates.json"
        with report_path.open("w", encoding="utf-8") as f:
            json.dump(suspect_pdf_duplicate_records, f, ensure_ascii=False, indent=2)
        _save_suspect_pdf_skip_keys(persistent_suspect_pdf_skip_keys | suspect_pdf_duplicate_keys)
        logger.error(
            "  Suspicious duplicate PDFs skipped: %s keys; report: %s",
            len(suspect_pdf_duplicate_keys),
            report_path,
        )
        print(f"  可疑重复 PDF 跳过: {len(suspect_pdf_duplicate_keys)} 篇")
        print(f"  可疑重复报告: {report_path}")
    if exact_pdf_duplicate_skipped:
        print(f"  完全重复 PDF 跳过: {exact_pdf_duplicate_skipped} 篇")

    parse_budget_stopped = False
    if max_parse_papers > 0 and len(need_parse) > max_parse_papers:
        original_need_parse = len(need_parse)
        need_parse = need_parse[:max_parse_papers]
        parse_budget_stopped = True
        logger.warning(
            "  Parse budget reached: processing %s of %s papers this run",
            len(need_parse),
            original_need_parse,
        )
        print(f"  解析预算: 本轮只解析 {len(need_parse)}/{original_need_parse} 篇")

    # -- Phase 2: Batch MinerU parsing --
    parsed: dict[str, tuple] = {}  # item_key -> (item, markdown_text)
    parse_failed_empty: set[str] = set()

    parse_api_limited = False

    if need_parse:
        print()
        print("=" * 60)
        print("  Phase 2: " + ("本地 PDF 文本解析" if local_parse_only else "MinerU 批量解析"))
        print("=" * 60)

        if local_parse_only:
            for item, pdf_path in need_parse:
                key = item["key"]
                if not _try_local_parse_fallback(item, pdf_path, parsed):
                    parse_failed_empty.add(key)
                    _record_parse_failure(parse_failures, key, "local_fallback_failed")
            need_parse = []

        http_client = httpx.Client(
            headers={"Authorization": f"Bearer {config.MINERU_TOKEN}"},
            timeout=httpx.Timeout(
                connect=30.0,
                read=config.MINERU_HTTP_TIMEOUT,
                write=config.MINERU_HTTP_TIMEOUT,
                pool=30.0,
            ),
        )
        api_limited = False
        try:
            for batch_start in range(0, len(need_parse), mineru_batch_size):
                batch_end = min(batch_start + mineru_batch_size, len(need_parse))
                batch = need_parse[batch_start:batch_end]
                batch_num = batch_start // mineru_batch_size + 1
                total_batches = (len(need_parse) + mineru_batch_size - 1) // mineru_batch_size

                print(f"\n  --- Batch {batch_num}/{total_batches} ({len(batch)} papers) ---")

                pdf_paths = [pdf_path for _, pdf_path in batch]
                item_keys = [item["key"] for item, _ in batch]

                try:
                    results = _submit_and_wait_batch(http_client, pdf_paths, item_keys)
                    for item, pdf_path in batch:
                        key = item["key"]
                        md = results.get(key)
                        if md and md.strip():
                            parsed[key] = (item, md)
                        else:
                            if not _try_local_parse_fallback(item, pdf_path, parsed):
                                parse_failed_empty.add(key)
                                _record_parse_failure(parse_failures, key, "mineru_failed_or_empty")
                except Exception as e:
                    if _is_api_limit_error(e):
                        logger.error(f"  MinerU API limit reached; stopping parse phase: {e}")
                        api_limited = True
                        parse_api_limited = True
                        break
                    logger.error(f"  Batch {batch_num} failed: {e}", exc_info=True)
                    # Fallback to per-paper parsing via pdf_parser (also httpx-based)
                    for item, pdf_path in batch:
                        key = item["key"]
                        try:
                            md = pdf_parser.parse_pdf(pdf_path, item_key=key, force=True)
                            if md and md.strip():
                                parsed[key] = (item, md)
                        except Exception as e2:
                            if _is_api_limit_error(e2):
                                logger.error(f"  MinerU API limit reached during fallback; stopping parse phase: {e2}")
                                api_limited = True
                                parse_api_limited = True
                                break
                            logger.error(f"  {key}: single-paper fallback also failed: {e2}")
                            if not _try_local_parse_fallback(item, pdf_path, parsed):
                                parse_failed_empty.add(key)
                                _record_parse_failure(parse_failures, key, "mineru_failed_or_empty")
                    if api_limited:
                        break

                # Brief rest between batches to avoid API rate limiting
                if batch_end < len(need_parse):
                    time.sleep(2)
        finally:
            http_client.close()

    print(f"\n  MinerU 解析完成: {len(parsed)} 篇")
    if parse_failed_empty:
        logger.warning(
            "  MinerU failed/empty outputs: %s keys: %s",
            len(parse_failed_empty),
            ", ".join(sorted(parse_failed_empty)[:30]),
        )

    # -- Phase 3: Chunking + Embedding + Ingestion --
    print()
    print("=" * 60)
    print("  Phase 3: 切块 + Embedding + 入库")
    print("=" * 60)

    # Merge cached + newly parsed results
    all_papers = {**cached, **parsed}

    stats = {
        "total": len(items),
        "daily_new_only": daily_new_only,
        "daily_since": _iso_utc(daily_since) if daily_since else None,
        "high_impact_only": high_impact_only,
        "min_impact_factor": min_impact_factor if high_impact_only else None,
        "require_q1": require_q1 if high_impact_only else None,
        "cached": len(cached),
        "parsed": len(parsed),
        "parse_failed_empty": len(parse_failed_empty),
        "parse_failed_empty_keys": sorted(parse_failed_empty),
        "no_pdf_skipped": no_pdf_skipped,
        "parse_failures_skipped": parse_failures_skipped,
        "parse_failure_max_attempts": config.PARSE_FAILURE_MAX_ATTEMPTS,
        "success": 0,
        "skipped": 0,
        "failed": 0,
        "chunks": 0,
        "api_limited": parse_api_limited,
        "parse_mode": "local" if local_parse_only else "mineru",
        "embed_provider": config.EMBED_PROVIDER,
        "embed_model": config.EMBED_MODEL,
        "parse_budget_stopped": parse_budget_stopped,
        "max_parse_papers": max_parse_papers or None,
        "embed_budget_stopped": False,
        "max_embed_papers": max_embed_papers or None,
        "max_embed_chunks": max_embed_chunks or None,
        "suspect_pdf_duplicates": len(suspect_pdf_duplicate_keys),
        "exact_pdf_duplicate_skipped": exact_pdf_duplicate_skipped,
    }
    stats["no_actionable"] = not all_papers and not need_parse and not stats["api_limited"]

    embedded_papers = 0
    for i, (key, (item, markdown_text)) in enumerate(all_papers.items(), 1):
        title = item.get("title", "?")[:50]
        cols = ", ".join(item.get("collection_names", []))
        print(f"\n[{i}/{len(all_papers)}] {title}")
        print(f"  Collection: {cols}")

        try:
            prepared_key, chunks, target_collections = _prepare_paper_chunks(item, markdown_text)
            planned_chunk_writes = len(chunks) * max(1, len(target_collections))
            if max_embed_papers > 0 and embedded_papers >= max_embed_papers:
                logger.warning("    Embedding paper budget reached (%s); stopping before API call", max_embed_papers)
                stats["embed_budget_stopped"] = True
                break
            if max_embed_chunks > 0 and planned_chunk_writes and stats["chunks"] + planned_chunk_writes > max_embed_chunks:
                logger.warning(
                    "    Embedding chunk budget reached (%s); next paper %s needs %s chunk writes, stopping before API call",
                    max_embed_chunks,
                    prepared_key,
                    planned_chunk_writes,
                )
                stats["embed_budget_stopped"] = True
                break

            n = _process_prepared_paper(prepared_key, chunks, target_collections)
            if n > 0:
                _clear_parse_failure(parse_failures, prepared_key)
                stats["success"] += 1
                stats["chunks"] += n
                embedded_papers += 1
            else:
                _record_parse_failure(parse_failures, prepared_key, "no_chunks")
                stats["skipped"] += 1
        except Exception as e:
            if _is_api_limit_error(e):
                logger.error(f"    API limit reached; stopping ingestion for today: {e}")
                stats["api_limited"] = True
                break
            logger.error(f"    FAILED: {e}", exc_info=True)
            stats["failed"] += 1

        # Rest every 5 papers to avoid Embedding API rate limiting
        if i % 5 == 0:
            time.sleep(0.5)

    # -- Statistics --
    print("\n" + "=" * 60)
    print("  入库统计")
    print(f"  总计:       {stats['total']} 篇")
    print(f"  缓存命中:   {stats['cached']} 篇")
    print(f"  解析模式:    {stats['parse_mode']}")
    print(f"  解析成功:    {stats['parsed']} 篇")
    print(f"  解析空结果:  {stats['parse_failed_empty']} 篇")
    print(f"  无 PDF 跳过: {stats['no_pdf_skipped']} 篇")
    print(f"  解析失败跳过: {stats['parse_failures_skipped']} 篇")
    print(f"  无可入库新增: {'是' if stats['no_actionable'] else '否'}")
    print(f"  入库成功:   {stats['success']} 篇")
    print(f"  跳过:       {stats['skipped']} 篇")
    print(f"  失败:       {stats['failed']} 篇")
    print(f"  API限制停止: {'是' if stats['api_limited'] else '否'}")
    print(f"  解析预算停止: {'是' if stats['parse_budget_stopped'] else '否'}")
    print(f"  预算停止:   {'是' if stats['embed_budget_stopped'] else '否'}")
    print(f"  新增块:     {stats['chunks']} 个")
    print("=" * 60)

    # ChromaDB status
    print("\nChromaDB Collections:")
    try:
        for col in vector_store.list_collections():
            print(f"  {col['name']} (safe={col['safe_name']}): {col['count']} chunks")
    except Exception as e:
        print(f"  (error: {e})")

    # Save statistics
    stats_path = _save_stats(stats, parse_failures)
    if daily_new_only and not stats["api_limited"]:
        _save_daily_state(run_started_at, stats)
    print(f"\nStats saved: {stats_path}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Zotero LLM Wiki Batch Ingest")
    parser.add_argument("--no-incremental", action="store_true", help="full ingest (ignore existing)")
    parser.add_argument("--incremental", action="store_true", help="only new papers (default)")
    parser.add_argument("--limit", type=int, default=0, help="max papers to process (0=all)")
    parser.add_argument("--force-parse", action="store_true", help="re-parse all PDFs")
    parser.add_argument("--collection", type=str, default=None, help="filter by collection name (fuzzy)")
    parser.add_argument("--batch-size", type=int, default=MINERU_BATCH_SIZE, help=f"MinerU batch size (default: {MINERU_BATCH_SIZE})")
    parser.add_argument("--high-impact-only", action="store_true", help="only ingest papers from high-impact JCR journals")
    parser.add_argument("--jcr-file", type=Path, default=DEFAULT_JCR_FILE, help="JCR xlsx path (or set ZOTERO_JCR_FILE)")
    parser.add_argument("--min-impact-factor", type=float, default=7.0, help="minimum Journal Impact Factor for --high-impact-only")
    parser.add_argument("--no-require-q1", action="store_true", help="allow non-Q1 journals if they pass min impact factor")
    parser.add_argument("--keep-unknown-journal", action="store_true", help="keep items whose journal cannot be matched in JCR")
    parser.add_argument("--dry-run", action="store_true", help="show selected candidates without parsing or embedding")
    parser.add_argument("--max-parse-papers", type=int, default=0, help="stop Phase 2 after this many papers that need MinerU parsing (0=unlimited)")
    parser.add_argument("--max-embed-papers", type=int, default=0, help="stop Phase 3 after this many successfully embedded papers (0=unlimited)")
    parser.add_argument("--max-embed-chunks", type=int, default=0, help="stop Phase 3 before exceeding this many chunk writes (0=unlimited)")
    parser.add_argument("--item-key", action="append", default=[], help="process only this Zotero item key; repeatable")
    parser.add_argument("--local-parse-only", action="store_true", help="use local PDF text extraction for parsing and do not call MinerU")
    parser.add_argument("--daily-new-only", action="store_true", help="process only Zotero items added after the previous daily run state")
    args = parser.parse_args()

    run(
        incremental=not args.no_incremental,
        limit=args.limit,
        force_parse=args.force_parse,
        collection_filter=args.collection,
        mineru_batch_size=args.batch_size,
        high_impact_only=args.high_impact_only,
        jcr_file=args.jcr_file,
        min_impact_factor=args.min_impact_factor,
        require_q1=not args.no_require_q1,
        keep_unknown_journal=args.keep_unknown_journal,
        dry_run=args.dry_run,
        max_parse_papers=args.max_parse_papers,
        max_embed_papers=args.max_embed_papers,
        max_embed_chunks=args.max_embed_chunks,
        item_keys=set(args.item_key) if args.item_key else None,
        local_parse_only=args.local_parse_only,
        daily_new_only=args.daily_new_only,
    )
